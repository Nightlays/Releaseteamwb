#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Доработка: исправлен «перенос людей» на следующий день.
# Теперь для первого снимка нового дня дневная дельта считается как:
#   (ws + retriedCount)_сейчас − (ws + retriedCount)_на конец предыдущего дня.
# А список people[день] формируется из логинов, у кого суммарная дневная дельта > 0
# (по cache["day_counts"][дата]), чтобы не попадали люди со вчера.
#
# В остальной логике НИЧЕГО не менялось.

import os, sys, json, time, base64, re
from datetime import datetime, timedelta, time as dtime
from zoneinfo import ZoneInfo
from typing import Dict, List, Set, Tuple, Optional
from io import BytesIO

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ===== Конфиг =====
BASE_URL   = "https://allure-testops.wb.ru"
PROJECT_ID = 7
API_TOKEN  = "c60f6235-440d-4657-983a-51dc71c53cf2"

# Google Sheets: ID таблицы
SPREADSHEET_ID = "17LfiixsqyrdJFps0Ek_W9nHYhtmZ0CgAFbYbTjK7iFg"

# Базовая папка в Drive, в ней создаём подпапку с именем релиза и уже туда пишем всё
PARENT_FOLDER_ID = "1Gc22V-oNEnZTJ0H-3T9GJ6PDEm82pT3H"

# Google Doc c логинами SWAT (строки вида: "login ФИО")
GDOC_SWAT_DOC_ID = "1AYHrg_w_aCdiunytlDVmbmyQAxiHRr51Z0Djju4GAnE"

# ===== Интервалы =====
OFFICIAL_SLOTS = [
    dtime(10,0), dtime(11,30), dtime(13,0), dtime(14,30),
    dtime(16,0), dtime(17,30), dtime(19,0), dtime(20,30), dtime(22,0)
]  # МСК

OUTPUT_BASENAME = "Статистика прохождения регрессов от SWAT (733)"

# ===== Стили =====
THIN = Side(style="thin", color="BDBDBD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(vertical="center", horizontal="center")
LEFT   = Alignment(vertical="center", horizontal="left")
GREEN_FILL   = PatternFill("solid", fgColor="C6EFCE")      # для интервалов с __done
ANDROID_FILL = PatternFill("solid", fgColor="FFB6D7A8")
IOS_FILL     = PatternFill("solid", fgColor="FFA4C2F4")
TOTAL_FILL   = PatternFill("solid", fgColor="FFFFE599")

# Цвета для "Итого кейсов за день" (по стриму)
TOT_GREEN = PatternFill("solid", fgColor="C8D77D")  # >30 или если последний интервал __done
TOT_20_29 = PatternFill("solid", fgColor="BFAF2A")  # 20..29
TOT_10_19 = PatternFill("solid", fgColor="D6B44A")  # 10..19
TOT_0_10  = PatternFill("solid", fgColor="E07C78")  # 0..10

# ===== Глобальные кеши =====
DRIVE_BASE_META: Dict[str, Optional[dict]] = {}   # защита от гонок по modifiedTime
DRIVE_VERSION_FOLDER: Dict[str, Optional[str]] = {}  # id подпапки версии

# ---------- Утилиты ----------
def now_msk() -> datetime:
    return datetime.now(ZoneInfo("Europe/Moscow"))

def cache_path(version: str) -> str:
    safe_version = re.sub(r"[^0-9A-Za-z\.\-_]", "_", version.strip())
    return f"cache_{safe_version}.json"

def b64_query(q: List[Dict]) -> str:
    return base64.b64encode(json.dumps(q, ensure_ascii=False).encode("utf-8")).decode("utf-8")

def build_session() -> requests.Session:
    if not API_TOKEN:
        print("[ERROR] Пустой API_TOKEN.")
        sys.exit(2)
    s = requests.Session()
    s.headers.update({
        "Accept":"application/json",
        "Authorization": "Api-Token " + API_TOKEN,
        "User-Agent":"swat-733-excel/1.6"
    })
    return s

def read_swat_logins_local(path: str) -> Set[str]:
    if not os.path.exists(path):
        return set()
    vals = set()
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            t = line.strip()
            if t:
                vals.add(t.lstrip("-•").strip().lower().split()[0])
    return vals

def ru_weekday_name(d: datetime) -> str:
    return ["Понедельник","Вторник","Среда","Четверг","Пятница","Суббота","Воскресенье"][d.weekday()]

def parse_time_label(lbl: str) -> Optional[dtime]:
    """Поддержка 'HH:MM' и 'HH:MM:SS' из старого кэша."""
    if isinstance(lbl, dtime):
        return lbl
    if isinstance(lbl, str):
        m = re.match(r"^(\d{2}):(\d{2})(?::\d{2})?$", lbl)
        if m:
            h, mnt = int(m.group(1)), int(m.group(2))
            return dtime(h, mnt, 0)
    return None

def time_to_label(t: dtime) -> str:
    return f"{t.hour:02d}:{t.minute:02d}"

# --- HTTP с ретраями ---
def http_get_with_retries(session: requests.Session, url: str, *, params: dict = None, timeout: int = 60):
    last_err = None
    for attempt in range(1, 4):
        try:
            r = session.get(url, params=params, timeout=timeout)
            r.raise_for_status()
            return r
        except Exception as e:
            last_err = e
            if attempt < 3:
                print(f"[WARN] GET {url} failed (attempt {attempt}/3): {e}. Retrying in 5s...")
                time.sleep(5)
            else:
                print(f"[ERROR] GET {url} failed after 3 attempts. params={params} err={e}")
                raise

# ---------- API ----------
def fetch_launches_hb(session: requests.Session, version: str) -> List[Dict]:
    terms = [f"[High/Blocker][DeployLab] Регресс {version}", f"[High/Blocker] Регресс {version}"]
    result_by_id = {}
    size = 1000
    for term in terms:
        search_b64 = b64_query([{"id":"name","type":"string","value": term}])
        page = 0
        while True:
            params = {"page":page,"size":size,"search":search_b64,"projectId":PROJECT_ID,"preview":"true","sort":"createdDate,desc"}
            r = http_get_with_retries(session, f"{BASE_URL}/api/launch", params=params, timeout=60)
            data = r.json() or {}
            content = data.get("content") or []
            if not content: break
            for it in content:
                try:
                    result_by_id[int(it.get("id"))] = it
                except: pass
            if len(content) < size: break
            page += 1
    launches = list(result_by_id.values())
    launches.sort(key=lambda L: (L.get("name",""), -int(L.get("id") or 0)))
    return launches

def fetch_member_stats(session: requests.Session, launch_id: int) -> List[Dict]:
    r = http_get_with_retries(session, f"{BASE_URL}/api/launch/{launch_id}/memberstats",
                              params={"size":1000,"page":0}, timeout=60)
    data = r.json() or []
    if isinstance(data, dict): return data.get("content") or []
    return data

def fetch_total_statistic(session: requests.Session, launch_id: int) -> List[Dict]:
    r = http_get_with_retries(session, f"{BASE_URL}/api/launch/{launch_id}/statistic", timeout=60)
    data = r.json() or []
    return data if isinstance(data, list) else []

def stat_total_count(stat_list: List[Dict]) -> int:
    return sum(int(x.get("count") or 0) for x in stat_list)

def has_inprogress_from_memberstats(memberstats: List[Dict]) -> bool:
    # есть «count» без статуса? тогда в работе
    no_status_sum = 0
    for m in memberstats:
        for it in (m.get("statistic") or []):
            if it.get("status") is None:
                try:
                    no_status_sum += int(it.get("count") or 0)
                except:
                    pass
    return no_status_sum > 0

def count_without_status(memberstats: List[Dict]) -> int:
    s = 0
    for m in memberstats:
        for it in (m.get("statistic") or []):
            if it.get("status") is None:
                try:
                    s += int(it.get("count") or 0)
                except:
                    pass
    return s

def is_launch_complete(total_stat: List[Dict], memberstats: List[Dict]) -> bool:
    non_final = {"unknown","queued","new","scheduled","pending"}
    for it in total_stat:
        key = str(it.get("status") or it.get("name") or it.get("key") or "").lower()
        cnt = int(it.get("count") or 0)
        if key in non_final and cnt > 0:
            return False
    if has_inprogress_from_memberstats(memberstats):
        return False
    return True

def swat_totals_by_person(memberstats: List[Dict], swat_set: Set[str]) -> Dict[str,int]:
    """
    Суммируем ТОЛЬКО элементы со статусом (failed/broken/passed/skipped/unknown и т.п.).
    'count' без 'status' — игнорируем.
    """
    per = {}
    for m in memberstats:
        assignee = (m.get("assignee") or "").strip().lower()
        if not assignee or assignee not in swat_set:
            continue
        s = m.get("statistic") or []
        cnt = 0
        for it in s:
            if "status" not in it or it.get("status") is None:
                continue  # игнорируем «без статуса»
            try:
                cnt += int(it.get("count") or 0)
            except:
                pass
        per[assignee] = per.get(assignee, 0) + cnt
    return per

def swat_retries_by_person(memberstats: List[Dict], swat_set: Set[str]) -> Dict[str,int]:
    """
    Количество ретраев по исполнителю (retriedCount). Нужен ТОЛЬКО для day_seen/day_counts.
    """
    per = {}
    for m in memberstats:
        assignee = (m.get("assignee") or "").strip().lower()
        if not assignee or assignee not in swat_set:
            continue
        try:
            rt = int(m.get("retriedCount") or 0)
        except:
            rt = 0
        per[assignee] = per.get(assignee, 0) + rt
    return per

# ---------- Google OAuth (Desktop) ----------
def _get_oauth_credentials(scopes: List[str]):
    from google.oauth2.credentials import Credentials as UserCredentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.auth.transport.requests import Request
    from google.auth.exceptions import RefreshError

    script_dir = os.path.dirname(os.path.abspath(__file__))
    token_path = os.path.join(script_dir, "token.json")
    client_secret_path = os.path.join(script_dir, "client_secret.json")

    creds = None
    if os.path.exists(token_path):
        try:
            creds = UserCredentials.from_authorized_user_file(token_path, scopes)
        except Exception:
            creds = None

    if creds and creds.valid:
        return creds

    if creds and creds.expired and creds.refresh_token:
        try:
            creds.refresh(Request())
            with open(token_path, "w", encoding="utf-8") as f:
                f.write(creds.to_json())
            return creds
        except RefreshError:
            try: os.replace(token_path, token_path + ".bak")
            except Exception: pass
            creds = None
        except Exception:
            try: os.replace(token_path, token_path + ".bak")
            except Exception: pass
            creds = None

    flow = InstalledAppFlow.from_client_secrets_file(
        client_secret_path,
        scopes=scopes
    )
    creds = flow.run_local_server(
        port=0,
        prompt="consent",
        access_type="offline"
    )

    with open(token_path, "w", encoding="utf-8") as f:
        f.write(creds.to_json())

    return creds

# ---------- Google Drive helpers ----------
def _get_drive_service():
    try:
        from googleapiclient.discovery import build
    except Exception as e:
        print(f"[WARN] google api libs not installed for Drive: {e}")
        return None
    scopes = ["https://www.googleapis.com/auth/drive"]
    creds = _get_oauth_credentials(scopes)
    return build("drive", "v3", credentials=creds)

def _drive_find_file(drive, folder_id: str, name: str) -> Optional[dict]:
    if not drive: return None
    q = f"name = '{name}' and '{folder_id}' in parents and trashed = false"
    try:
        resp = drive.files().list(
            q=q,
            fields="files(id,name,parents,modifiedTime,md5Checksum,mimeType)",
            pageSize=1,
            supportsAllDrives=True,
            includeItemsFromAllDrives=True
        ).execute()
        files = resp.get("files", [])
        return files[0] if files else None
    except Exception as e:
        print(f"[WARN] Drive list failed: {e}")
        return None

def _drive_get_version_folder(version: str, create_if_missing: bool) -> Optional[str]:
    if version in DRIVE_VERSION_FOLDER and DRIVE_VERSION_FOLDER[version]:
        return DRIVE_VERSION_FOLDER[version]
    drive = _get_drive_service()
    if not drive:
        return None
    q = f"mimeType = 'application/vnd.google-apps.folder' and name = '{version}' and '{PARENT_FOLDER_ID}' in parents and trashed = false"
    try:
        resp = drive.files().list(
            q=q,
            fields="files(id,name,parents)",
            pageSize=1,
            supportsAllDrives=True,
            includeItemsFromAllDrives=True
        ).execute()
        files = resp.get("files", [])
        if files:
            fid = files[0]["id"]
            DRIVE_VERSION_FOLDER[version] = fid
            return fid
        if not create_if_missing:
            return None
        meta = {
            "name": version,
            "mimeType": "application/vnd.google-apps.folder",
            "parents": [PARENT_FOLDER_ID]
        }
        created = drive.files().create(
            body=meta,
            fields="id",
            supportsAllDrives=True
        ).execute()
        fid = created.get("id")
        DRIVE_VERSION_FOLDER[version] = fid
        return fid
    except Exception as e:
        print(f"[WARN] Drive create/get version folder failed: {e}")
        return None

def _drive_download_cache(version: str) -> Tuple[Optional[dict], Optional[dict]]:
    drive = _get_drive_service()
    if not drive:
        return None, None
    folder_id = _drive_get_version_folder(version, create_if_missing=False)
    if not folder_id:
        return None, None
    name = f"cache_{version}.json"
    meta = _drive_find_file(drive, folder_id, name)
    if not meta:
        return None, None
    try:
        from googleapiclient.http import MediaIoBaseDownload
        req = drive.files().get_media(fileId=meta["id"])
        buf = BytesIO()
        downloader = MediaIoBaseDownload(buf, req)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        data = json.loads(buf.getvalue().decode("utf-8"))
        return data, meta
    except Exception as e:
        print(f"[WARN] Drive download failed: {e}")
        return None, None

def _merge_maps_sum(A: dict, B: dict) -> dict:
    R = dict(A or {})
    for k, v in (B or {}).items():
        try:
            v = int(v)
        except:
            continue
        if k in R:
            R[k] = max(int(R[k]), v)
        else:
            R[k] = v
    return R

def _merge_caches(A: dict, B: dict) -> dict:
    R = {"runs": {}, "history": {}, "people": {}}

    # runs
    keys_runs = set((A.get("runs") or {}).keys()) | set((B.get("runs") or {}).keys())
    for lid in keys_runs:
        ra = (A.get("runs") or {}).get(lid) or {}
        rb = (B.get("runs") or {}).get(lid) or {}
        R["runs"][lid] = {
            "name":      rb.get("name")      or ra.get("name"),
            "platform":  rb.get("platform")  or ra.get("platform"),
            "total_cases": max(int(ra.get("total_cases") or 0), int(rb.get("total_cases") or 0)),
            "last_total":  max(int(ra.get("last_total")  or 0), int(rb.get("last_total")  or 0)),
        }

    # history
    keys_days = set((A.get("history") or {}).keys()) | set((B.get("history") or {}).keys())
    for day in keys_days:
        dayA = (A.get("history") or {}).get(day) or {}
        dayB = (B.get("history") or {}).get(day) or {}
        R["history"][day] = {}
        for lid in set(dayA.keys()) | set(dayB.keys()):
            lidA = dayA.get(lid) or {}
            lidB = dayB.get(lid) or {}
            merged = {}
            for k in set(lidA.keys()) | set(lidB.keys()):
                va, vb = lidA.get(k), lidB.get(k)
                if k.endswith("__done"):
                    merged[k] = bool(va) or bool(vb)
                elif k.endswith("__rem") or re.match(r"^\d{2}:\d{2}(:\d{2})?$", k or ""):
                    merged[k] = max(int(va or 0), int(vb or 0))
                else:
                    merged[k] = vb if vb is not None else va
            R["history"][day][lid] = merged

    # people
    keys_people = set((A.get("people") or {}).keys()) | set((B.get("people") or {}).keys())
    for day in keys_people:
        sa = set(A.get("people", {}).get(day) or [])
        sb = set(B.get("people", {}).get(day) or [])
        R["people"][day] = sorted(sa | sb)

    # === NEW: day_counts (опционально) ===
    R["day_counts"] = {}
    keys_dc = set((A.get("day_counts") or {}).keys()) | set((B.get("day_counts") or {}).keys())
    for day in keys_dc:
        R["day_counts"][day] = _merge_maps_sum(A.get("day_counts", {}).get(day) or {},
                                               B.get("day_counts", {}).get(day) or {})

    # === NEW: day_seen (опционально) ===
    R["day_seen"] = {}
    keys_ds = set((A.get("day_seen") or {}).keys()) | set((B.get("day_seen") or {}).keys())
    for day in keys_ds:
        R["day_seen"][day] = _merge_maps_sum(A.get("day_seen", {}).get(day) or {},
                                             B.get("day_seen", {}).get(day) or {})

    return R

def _drive_upload_cache(version: str, data: dict):
    drive = _get_drive_service()
    if not drive:
        return
    folder_id = _drive_get_version_folder(version, create_if_missing=True)
    if not folder_id:
        print("[WARN] No Drive folder for version; skip Drive cache upload.")
        return
    name = f"cache_{version}.json"
    try:
        from googleapiclient.http import MediaIoBaseUpload, MediaIoBaseDownload
    except Exception as e:
        print(f"[WARN] google api libs not installed for Drive upload: {e}")
        return

    current = _drive_find_file(drive, folder_id, name)
    base_meta = DRIVE_BASE_META.get(version)

    payload = json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")
    media = MediaIoBaseUpload(BytesIO(payload), mimetype="application/json", resumable=False)

    try:
        if current:
            if base_meta and current.get("modifiedTime") != base_meta.get("modifiedTime"):
                try:
                    req = drive.files().get_media(fileId=current["id"])
                    buf = BytesIO()
                    downloader = MediaIoBaseDownload(buf, req)
                    done = False
                    while not done:
                        _, done = downloader.next_chunk()
                    remote = json.loads(buf.getvalue().decode("utf-8"))
                    data = _merge_caches(data, remote)
                    payload = json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")
                    media = MediaIoBaseUpload(BytesIO(payload), mimetype="application/json", resumable=False)
                except Exception as me:
                    print(f"[WARN] conflict merge failed, proceed with local: {me}")

            drive.files().update(
                fileId=current["id"],
                media_body=media,
                supportsAllDrives=True
            ).execute()
        else:
            file_meta = {"name": name, "parents": [folder_id], "mimeType": "application/json"}
            drive.files().create(
                body=file_meta,
                media_body=media,
                fields="id",
                supportsAllDrives=True
            ).execute()

        # Снапшот
        snap_name = f"cache_{version}_{now_msk().strftime('%Y%m%d-%H%M')}.json"
        snap_meta = {"name": snap_name, "parents": [folder_id], "mimeType": "application/json"}
        drive.files().create(
            body=snap_meta,
            media_body=MediaIoBaseUpload(BytesIO(payload), mimetype="application/json", resumable=False),
            fields="id",
            supportsAllDrives=True
        ).execute()
    except Exception as e:
        print(f"[WARN] Drive upload failed: {e}")

# ---------- SWAT из Google Doc ----------
def _get_drive_export_text(file_id: str) -> Optional[str]:
    drive = _get_drive_service()
    if not drive:
        return None
    try:
        data = drive.files().export(fileId=file_id, mimeType="text/plain").execute()
        return data.decode("utf-8") if isinstance(data, (bytes, bytearray)) else str(data)
    except Exception:
        return None

def read_swat_logins_from_gdoc() -> Set[str]:
    text = _get_drive_export_text(GDOC_SWAT_DOC_ID)
    if text is None:
        print("[INFO] Не удалось получить SWAT из Google Doc. Использую локальный SWAT.txt")
        script_dir = os.path.dirname(os.path.abspath(__file__))
        return read_swat_logins_local(os.path.join(script_dir, "SWAT.txt"))
    logins: Set[str] = set()
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        first = line.split()[0].strip().lower()
        if first and not first.startswith("#"):
            logins.add(first)
    if not logins:
        print("[INFO] SWAT в Google Doc пустой. Использую локальный SWAT.txt")
        script_dir = os.path.dirname(os.path.abspath(__file__))
        return read_swat_logins_local(os.path.join(script_dir, "SWAT.txt"))
    return logins

# ---------- Кэш ----------
def load_cache(version: str) -> Dict:
    data, meta = _drive_download_cache(version)
    if data is not None:
        DRIVE_BASE_META[version] = meta
        data.setdefault("day_counts", {})
        data.setdefault("day_seen", {})
        return data

    path = cache_path(version)
    DRIVE_BASE_META[version] = None
    if os.path.exists(path):
        try:
            obj = json.load(open(path, "r", encoding="utf-8"))
            obj.setdefault("runs", {})
            obj.setdefault("history", {})
            obj.setdefault("people", {})
            obj.setdefault("day_counts", {})  # NEW optional
            obj.setdefault("day_seen", {})    # NEW optional
            return obj
        except:
            pass
    return {"runs":{}, "history":{}, "people":{}, "day_counts":{}, "day_seen":{}}

def save_cache(version: str, obj: Dict):
    path = cache_path(version)
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=2)
    os.replace(tmp, path)
    _drive_upload_cache(version, obj)

# ---------- Excel ----------
def auto_fit(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 6
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, col).value
            if v is None: continue
            max_len = max(max_len, len(str(v)))
        width = min(max(max_len + 2, 12), 48)
        ws.column_dimensions[get_column_letter(col)].width = width

def write_day_header(ws, start_row: int, day_label: str, slot_list: List[dtime], ios_start_col: int) -> Tuple[int,int]:
    ws.cell(start_row+0, 1, day_label).font = Font(bold=True)

    ws.merge_cells(start_row=start_row+3, start_column=2, end_row=start_row+3, end_column=1+len(slot_list))
    ws.cell(start_row+3, 2, "Количество кейсов по временным промежуткам ").font = Font(bold=True)
    ws.cell(start_row+3, 2).alignment = CENTER
    ws.merge_cells(start_row=start_row+3, start_column=ios_start_col, end_row=start_row+3, end_column=ios_start_col+len(slot_list)-1)
    ws.cell(start_row+3, ios_start_col, "Количество кейсов по временным промежуткам ").font = Font(bold=True)
    ws.cell(start_row+3, ios_start_col).alignment = CENTER

    ws.cell(start_row+4, 1, "Стрим").font = Font(bold=True)
    ws.cell(start_row+4,1).fill = ANDROID_FILL
    for i, t in enumerate(slot_list):
        ws.cell(start_row+4, 2+i, t)
        ws.cell(start_row+4, 2+i).fill = ANDROID_FILL
    a_day_total_col = 2 + len(slot_list)
    ws.cell(start_row+4, a_day_total_col, "Итого кейсов за день").font = Font(bold=True)

    ws.cell(start_row+4, ios_start_col-1, "Стрим").font = Font(bold=True)
    ws.cell(start_row+4, ios_start_col-1).fill = IOS_FILL
    for i, t in enumerate(slot_list):
        ws.cell(start_row+4, ios_start_col+i, t)
        ws.cell(start_row+4, ios_start_col+i).fill = IOS_FILL
    i_day_total_col = ios_start_col + len(slot_list)
    ws.cell(start_row+4, i_day_total_col, "Итого кейсов за день").font = Font(bold=True)

    for r in range(start_row, start_row+5):
        for c in range(1, i_day_total_col+1+5):
            ws.cell(r,c).alignment = CENTER
            ws.cell(r,c).border = BORDER

    return a_day_total_col, i_day_total_col

def short_stream_name(fullname: str) -> str:
    m = re.search(r"\[Stream\s+([^\]]+)\]", fullname)
    if m:
        return m.group(1)
    return fullname

def build_full_table(runs: Dict[int, Dict], history: Dict[str, Dict[str, Dict[str, int]]],
                     people: Dict[str, List[str]], rel_version: str, now_dt: datetime,
                     day_counts: Dict[str, Dict[str, int]]) -> Tuple[Workbook, str]:
    android = [(lid, info) for lid, info in runs.items() if info.get("platform") == "Android"]
    ios     = [(lid, info) for lid, info in runs.items() if info.get("platform") == "iOS"]
    android.sort(key=lambda x: x[1]["name"])
    ios.sort(key=lambda x: x[1]["name"])

    wb = Workbook()
    ws = wb.active
    ws.title = rel_version

    date_keys = sorted(history.keys()) if history else []

    global_slots = set( OFFICIAL_SLOTS )
    for date_str in date_keys:
        for _, slotmap in (history.get(date_str, {}) or {}).items():
            for k in slotmap.keys():
                if k.endswith("__done") or k.endswith("__rem"): continue
                t = parse_time_label(str(k))
                if t:
                    global_slots.add(t)
    slot_list_global = sorted(list(global_slots))

    ios_start_col_const = max(11, 2 + len(slot_list_global) + 2)

    ws.cell(1, 1, "Android").font = Font(bold=True, size=12)
    ws.cell(1, ios_start_col_const-1, "iOS").font = Font(bold=True, size=12)
    ws.cell(1, 7, "Итого кейсов").font = Font(bold=True); ws.cell(1,7).fill = TOTAL_FILL
    ws.cell(1, 17, "Итого кейсов").font = Font(bold=True); ws.cell(1,17).fill = TOTAL_FILL
    ws.cell(3, 1, "Blocker + High").font = Font(bold=True)

    last_total_by_date: Dict[str, Dict[int, int]] = {}
    for date_str in date_keys:
        last_total_by_date[date_str] = {}
        day_map = history.get(date_str, {}) or {}
        for lid_str, slots_map in day_map.items():
            lid = int(lid_str)
            last_val = 0
            for k, v in slots_map.items():
                if k.endswith("__done") or k.endswith("__rem"): continue
                t = parse_time_label(str(k))
                if t:
                    try:
                        last_val = max(last_val, int(v))
                    except:
                        pass
            last_total_by_date[date_str][lid] = last_val

    start_row = 4
    max_rows = max(len(android), len(ios))
    header_rows = 5
    spacer_rows = 3

    total_android_all_days = 0
    total_ios_all_days = 0

    for idx_date, date_str in enumerate(date_keys):
        d = datetime.strptime(date_str, "%Y-%m-%d")
        day_label = f"{ru_weekday_name(d)} {d.strftime('%d.%m')}"

        a_day_total_col, i_day_total_col = write_day_header(ws, start_row, day_label, slot_list_global, ios_start_col_const)

        per_day = {"Android": {}, "iOS": {}}
        launches = history.get(date_str, {})
        for lid_str, slotmap in launches.items():
            lid = int(lid_str)
            info = runs.get(lid)
            if not info:
                continue
            plat = info["platform"]
            if plat not in per_day:  # пропускаем «Other», чтобы не ловить KeyError
                continue
            agg = {t:0 for t in slot_list_global}
            for k, v in slotmap.items():
                if k.endswith("__done") or k.endswith("__rem"): continue
                t = parse_time_label(str(k))
                if t and t in agg:
                    try: agg[t] = int(v)
                    except: pass
            per_day[plat][lid] = agg

        row_start = start_row + header_rows
        row = row_start
        row_count = max_rows

        def day_delta_for_launch(lid: int) -> int:
            cur = int(last_total_by_date.get(date_str, {}).get(lid, 0))
            if idx_date == 0:
                return cur
            prev_date = date_keys[idx_date-1]
            prev = int(last_total_by_date.get(prev_date, {}).get(lid, 0))
            return max(0, cur - prev)

        android_day_sum = 0
        ios_day_sum = 0

        def write_platform_row(start_col, lid, info, slot_vals, delta_val, is_ios=False):
            name_cell = ws.cell(row, start_col-1, short_stream_name(info["name"]))
            name_cell.alignment = LEFT
            name_cell.hyperlink = f"{BASE_URL}/launch/{lid}"
            name_cell.style = "Hyperlink"

            for i, t in enumerate(slot_list_global):
                v = int(slot_vals.get(t, 0))
                ws.cell(row, start_col+i, v).number_format = "0"
                ws.cell(row, start_col+i).alignment = CENTER
            ws.cell(row, start_col+len(slot_list_global), delta_val).number_format = "0"
            ws.cell(row, start_col+len(slot_list_global)).alignment = CENTER
            for c in range(start_col-1, start_col+len(slot_list_global)+1):
                ws.cell(row, c).border = BORDER

            lid_map = (history.get(date_str, {}) or {}).get(str(lid), {}) or {}
            last_label = None
            last_index = None
            for idx_t, t in reversed(list(enumerate(slot_list_global))):
                norm = time_to_label(t)
                alt  = f"{norm}:00"
                if norm in lid_map or alt in lid_map:
                    last_label = norm if norm in lid_map else alt
                    last_index = idx_t
                    break

            for i, t in enumerate(slot_list_global):
                norm = time_to_label(t)
                alt = f"{norm}:00" if False else f"{norm}:00"
                cell = ws.cell(row, start_col + i)
                if lid_map.get(f"{norm}__done") or lid_map.get(f"{alt}__done"):
                    cell.fill = GREEN_FILL

            if last_label is not None and last_index is not None:
                rem = int(lid_map.get(f"{last_label}__rem", 0))
                cell_last = ws.cell(row, start_col + last_index)
                v_last = cell_last.value
                if rem > 0:
                    cell_last.value = f"{v_last} ({rem})"
                    cell_last.alignment = CENTER

            total_cell = ws.cell(row, start_col + len(slot_list_global))
            done_flag = False
            if last_label:
                if lid_map.get(f"{last_label}__done"):
                    done_flag = True
            if done_flag:
                total_cell.fill = TOT_GREEN
            else:
                if delta_val > 30:
                    total_cell.fill = TOT_GREEN
                elif 20 <= delta_val <= 30:
                    total_cell.fill = TOT_20_29
                elif 10 <= delta_val <= 19:
                    total_cell.fill = TOT_10_19
                else:
                    total_cell.fill = TOT_0_10

        for idx in range(row_count):
            if idx < len(android):
                lid, info = android[idx]
                vals = per_day.get("Android", {}).get(lid, {t:0 for t in slot_list_global})
                delta_val = day_delta_for_launch(lid)
                android_day_sum += delta_val
                write_platform_row(2, lid, info, vals, delta_val, is_ios=False)

            if idx < len(ios):
                lid2, info2 = ios[idx]
                vals2 = per_day.get("iOS", {}).get(lid2, {t:0 for t in slot_list_global})
                delta_val2 = day_delta_for_launch(lid2)
                ios_day_sum += delta_val2
                write_platform_row(ios_start_col_const, lid2, info2, vals2, delta_val2, is_ios=True)

            row += 1

        day_cases = android_day_sum + ios_day_sum

        # ====== число сотрудников за день по day_counts ======
        active_count = 0
        if date_str in (day_counts or {}):
            try:
                active_count = sum(1 for v in (day_counts[date_str] or {}).values() if int(v) > 0)
            except:
                active_count = 0

        wd = d.weekday()  # Mon=0 ... Sun=6
        if wd == 4:            # Friday
            minutes_per_person_fixed = 0
        elif wd in (5, 6, 0):  # Saturday, Sunday, Monday
            minutes_per_person_fixed = 540
        elif wd == 1:          # Tuesday
            minutes_per_person_fixed = 0
        else:
            minutes_per_person_fixed = 0

        metrics_col = (a_day_total_col + ios_start_col_const) // 2
        labels_row  = start_row + 1
        values_row  = start_row + 2

        ws.cell(labels_row, metrics_col + 0, "Кол-во кейсов за день:").alignment = CENTER
        ws.cell(labels_row, metrics_col + 1, "Сотрудников:").alignment = CENTER
        ws.cell(labels_row, metrics_col + 2, "Мин/чел:").alignment = CENTER
        ws.cell(labels_row, metrics_col + 3, "Средн. мин/кейс:").alignment = CENTER

        ws.cell(values_row, metrics_col + 0, int(day_cases))
        ws.cell(values_row, metrics_col + 1, int(active_count))
        ws.cell(values_row, metrics_col + 2, int(minutes_per_person_fixed))

        A = get_column_letter(metrics_col + 0)
        B = get_column_letter(metrics_col + 1)
        C = get_column_letter(metrics_col + 2)
        avg_formula = f'=IFERROR({C}{values_row}*{B}{values_row}/{A}{values_row},0)'
        avg_cell = ws.cell(values_row, metrics_col + 3)
        avg_cell.value = avg_formula
        avg_cell.number_format = "0.0"

        for c in range(metrics_col, metrics_col + 4):
            ws.cell(labels_row, c).font = Font(bold=True)
            ws.cell(labels_row, c).border = BORDER
            ws.cell(values_row, c).alignment = CENTER
            ws.cell(values_row, c).font = Font(bold=True)
            ws.cell(values_row, c).border = BORDER

        total_android_all_days += android_day_sum
        total_ios_all_days += ios_day_sum

        start_row = start_row + header_rows + max_rows + spacer_rows

    ws.cell(2, 7, int(total_android_all_days)).alignment = CENTER
    ws.cell(2, 17, int(total_ios_all_days)).alignment = CENTER

    auto_fit(ws)
    wb_name = f"Прохождение SWAT {rel_version}.xlsx"
    return wb, wb_name

# ---------- Google Sheets upload ----------
def upload_excel_sheet_to_gsheets(excel_path: str, spreadsheet_id: str, sheet_name: str, version: str) -> bool:
    script_dir = os.path.dirname(os.path.abspath(__file__))
    log_path = os.path.join(script_dir, "gsheets_upload.log")
    def _log(msg: str):
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(f"[{now_msk().strftime('%Y-%m-%d %H:%M:%S')}] {msg}\n")

    try:
        try:
            from googleapiclient.discovery import build  # noqa
            from googleapiclient.http import MediaFileUpload  # noqa
            from googleapiclient.errors import HttpError  # noqa
        except Exception as e:
            _log(f"ERROR: google api libs not installed: {e}")
            return False

        from googleapiclient.discovery import build
        from googleapiclient.http import MediaFileUpload
        from googleapiclient.errors import HttpError

        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive"
        ]
        creds = _get_oauth_credentials(scopes)
        sheets_srv = build("sheets", "v4", credentials=creds)
        drive_srv  = build("drive",  "v3", credentials=creds)

        version_folder_id = _drive_get_version_folder(version, create_if_missing=True)
        if not version_folder_id:
            _log("WARN: Не удалось получить/создать подпапку версии. Пишу во вложения базовой папки.")
            version_folder_id = PARENT_FOLDER_ID

        wb = load_workbook(excel_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            _log(f"ERROR: В Excel нет листа '{sheet_name}'")
            return False
        ws = wb[sheet_name]
        max_row = ws.max_row
        max_col = ws.max_column
        values: List[List[str]] = []
        for r in range(1, max_row+1):
            row = []
            for c in range(1, max_col+1):
                v = ws.cell(r, c).value
                if isinstance(v, (datetime,)):
                    v = v.strftime("%Y-%m-%d %H:%M:%S")
                elif isinstance(v, dtime):
                    v = f"{v.hour:02d}:{v.minute:02d}"
                row.append("" if v is None else v)
            values.append(row)

        temp_spreadsheet_id = None
        try:
            _log(f"INFO: Using version folder {version_folder_id} for temp file.")
            media = MediaFileUpload(
                excel_path,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                resumable=False
            )
            file_meta = {
                "name": os.path.basename(excel_path),
                "mimeType": "application/vnd.google-apps.spreadsheet",
                "parents": [version_folder_id]
            }
            temp_file = drive_srv.files().create(
                body=file_meta, media_body=media, fields="id, parents", supportsAllDrives=True
            ).execute()
            temp_spreadsheet_id = temp_file.get("id")
            _log(f"INFO: Создана временная таблица {temp_spreadsheet_id}, parents={temp_file.get('parents')}")
        except HttpError as e:
            msg = str(e)
            _log(f"WARN: create temp failed: {msg}. Fallback to values.update.")
            try:
                target_meta = sheets_srv.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
                target_sheets = target_meta.get("sheets", [])
                target_sheet_id = None
                for s in target_sheets:
                    if s.get("properties", {}).get("title") == sheet_name:
                        target_sheet_id = s.get("properties", {}).get("sheetId")
                        break
                if target_sheet_id is None:
                    add_req = {"requests": [{"addSheet": {"properties": {"title": sheet_name}}}]}
                    sheets_srv.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_id, body=add_req).execute()
                    _log(f"INFO: Добавлен лист '{sheet_name}' (fallback)")
                else:
                    sheets_srv.spreadsheets().values().clear(
                        spreadsheetId=spreadsheet_id,
                        range=f"'{sheet_name}'!A:ZZZ"
                    ).execute()
                    _log(f"INFO: Очищен лист '{sheet_name}' (fallback)")

                sheets_srv.spreadsheets().values().update(
                    spreadsheetId=spreadsheet_id,
                    range=f"'{sheet_name}'!A1",
                    valueInputOption="RAW",
                    body={"values": values}
                ).execute()
                _log(f"SUCCESS: Данные записаны в лист '{sheet_name}' (fallback, без форматирования).")
                return True
            except Exception as fe:
                _log(f"ERROR: Fallback не удался: {fe}")
                return False

        try:
            temp_meta = sheets_srv.spreadsheets().get(spreadsheetId=temp_spreadsheet_id).execute()
            temp_sheets = temp_meta.get("sheets", [])
            source_sheet_id = None
            for s in temp_sheets:
                if s.get("properties", {}).get("title") == sheet_name:
                    source_sheet_id = s.get("properties", {}).get("sheetId")
                    break
            if source_sheet_id is None and temp_sheets:
                source_sheet_id = temp_sheets[0].get("properties", {}).get("sheetId")
                _log(f"WARN: Временная таблица не содержит лист '{sheet_name}', использую первый лист id={source_sheet_id}")
            if source_sheet_id is None:
                _log("ERROR: Не удалось определить исходный лист во временной таблице")
                return False

            target_meta = sheets_srv.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
            target_sheets = target_meta.get("sheets", [])
            to_delete_id = None
            for s in target_sheets:
                if s.get("properties", {}).get("title") == sheet_name:
                    to_delete_id = s.get("properties", {}).get("sheetId")
                    break
            if to_delete_id is not None:
                req = {"requests": [{"deleteSheet": {"sheetId": to_delete_id}}]}
                sheets_srv.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_id, body=req).execute()

            copy_result = sheets_srv.spreadsheets().sheets().copyTo(
                spreadsheetId=temp_spreadsheet_id,
                sheetId=source_sheet_id,
                body={"destinationSpreadsheetId": spreadsheet_id}
            ).execute()
            new_sheet_id = copy_result.get("sheetId")

            req = {"requests": [{
                "updateSheetProperties": {
                    "properties": {"sheetId": new_sheet_id, "title": sheet_name},
                    "fields": "title"
                }
            }]}
            sheets_srv.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_id, body=req).execute()
            _log(f"SUCCESS: Лист '{sheet_name}' обновлён (с форматированием).")
            return True

        finally:
            if temp_spreadsheet_id:
                try:
                    drive_srv.files().delete(fileId=temp_spreadsheet_id).execute()
                except Exception:
                    pass

    except Exception as e:
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(f"[{now_msk().strftime('%Y-%m-%d %H:%M:%S')}] ERROR: {e}\n")
        return False

# ---------- Маппинг ручного слота ----------
def pick_manual_slot_label(now_dt: datetime) -> Optional[str]:
    cur = now_dt.time()
    candidates = [t for t in sorted(OFFICIAL_SLOTS) if t <= cur]
    if not candidates:
        return None
    return time_to_label(candidates[-1])

# ---------- Вспомогательное: найти предыдущий день ----------
def _find_prev_day_key(day_seen_map: Dict[str, Dict[str,int]], current_key: str) -> Optional[str]:
    if not day_seen_map:
        return None
    prev_keys = [k for k in day_seen_map.keys() if k < current_key]
    if not prev_keys:
        return None
    return max(prev_keys)

# ---------- Сбор ----------
def run_collect(version: str, manual_between: bool):
    session = build_session()

    # SWAT-логины
    swat_set = read_swat_logins_from_gdoc()

    cache = load_cache(version)
    cache.setdefault("day_counts", {})  # опциональные разделы на всякий случай
    cache.setdefault("day_seen", {})

    now = now_msk()
    date_key = now.strftime("%Y-%m-%d")

    launches = fetch_launches_hb(session, version)

    # ---- НОВОЕ: готовим базовый снимок на границе суток ----
    prev_seen_same_day: Dict[str, int] = (cache.get("day_seen") or {}).get(date_key, {}) or {}
    prev_day_key = _find_prev_day_key(cache.get("day_seen") or {}, date_key)
    prev_seen_prev_day: Dict[str, int] = (cache.get("day_seen") or {}).get(prev_day_key, {}) or {}
    # база: если уже есть снимок сегодня — берём его; иначе берём снимок вчера (последнего дня перед текущим)
    base_seen_for_delta: Dict[str, int] = prev_seen_same_day if prev_seen_same_day else prev_seen_prev_day

    # ---- Текущий снимок ws+retry для day_seen/day_counts ----
    ws_cur: Dict[str, int] = {}

    for L in launches:
        lid = int(L.get("id"))
        name = L.get("name","")
        platform = "Android" if "[Android]" in name else ("iOS" if "[iOS]" in name else "Other")

        total_stat = fetch_total_statistic(session, lid)
        memberstats = fetch_member_stats(session, lid)
        total_cases = stat_total_count(total_stat)
        complete = is_launch_complete(total_stat, memberstats)
        rem_no_status = count_without_status(memberstats)

        per_person_ws = swat_totals_by_person(memberstats, swat_set)   # ТОЛЬКО статусы
        per_person_rt = swat_retries_by_person(memberstats, swat_set)  # retriedCount

        # НОВОЕ: ws_cur = ws + retriedCount (для day_seen/day_counts ТОЛЬКО)
        all_logins = set(per_person_ws.keys()) | set(per_person_rt.keys())
        for login in all_logins:
            ws = int(per_person_ws.get(login, 0))
            rt = int(per_person_rt.get(login, 0))
            ws_cur[login] = ws_cur.get(login, 0) + (ws + rt)

        # ===== ВЕСЬ ОСТАЛЬНЫЙ БЛОК НЕ МЕНЯЕМ =====
        runinfo = cache["runs"].setdefault(str(lid), {"name": name, "platform": platform, "last_total": 0, "total_cases": 0})
        runinfo["name"] = name
        runinfo["platform"] = platform
        runinfo["total_cases"] = total_cases
        runinfo["last_total"] = int(sum(per_person_ws.values()))  # last_total — ТОЛЬКО ws

        # history — пишем только ws (как и раньше)
        write_slot = None
        if manual_between:
            mapped = pick_manual_slot_label(now)
            if mapped:
                write_slot = mapped
        else:
            cur_label = time_to_label(dtime(now.hour, now.minute))
            if any(cur_label == time_to_label(t) for t in OFFICIAL_SLOTS):
                write_slot = cur_label

        if write_slot:
            day_map = cache["history"].setdefault(date_key, {})
            lid_map = day_map.setdefault(str(lid), {})
            lid_map[write_slot] = int(sum(per_person_ws.values()))
            lid_map[f"{write_slot}__rem"] = int(rem_no_status)
            if complete:
                lid_map[f"{write_slot}__done"] = True

    # ---- НОВОЕ: day_counts/day_seen и people по дельте дня ----
    # базовая карта для дельт:
    base_seen = {k:int(v) for k,v in (base_seen_for_delta or {}).items()}
    # текущий снимок:
    cur_seen  = {k:int(v) for k,v in (ws_cur or {}).items()}

    # предыдущие накопленные day_counts за сегодня (если уже были)
    day_counts_today: Dict[str, int] = (cache.get("day_counts") or {}).get(date_key, {}) or {}
    updated_day_counts: Dict[str, int] = dict(day_counts_today)

    all_logins_today = set(base_seen.keys()) | set(cur_seen.keys())
    for login in all_logins_today:
        cur = int(cur_seen.get(login, 0))
        prev = int(base_seen.get(login, 0))
        delta = cur - prev
        if delta > 0:
            updated_day_counts[login] = int(updated_day_counts.get(login, 0)) + delta
        else:
            updated_day_counts.setdefault(login, updated_day_counts.get(login, 0))

    cache.setdefault("day_counts", {})[date_key] = updated_day_counts
    cache.setdefault("day_seen", {})[date_key] = cur_seen

    # people[дата] = только логины с суммарной дневной дельтой > 0
    active_people_today = sorted([login for login, total in updated_day_counts.items() if int(total) > 0])
    cache.setdefault("people", {})[date_key] = active_people_today

    save_cache(version, cache)

    runs_int = {int(k): v for k, v in cache["runs"].items()}
    wb, fname = build_full_table(
        runs_int,
        cache.get("history", {}),
        cache.get("people", {}),
        version,
        now_msk(),
        cache.get("day_counts", {})
    )

    out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), fname)
    wb.save(out_path)
    print(f"[OK] Excel сохранён: {out_path}")

    ok = upload_excel_sheet_to_gsheets(out_path, SPREADSHEET_ID, version, version)
    if ok:
        print(f"[OK] Лист '{version}' загружен в Google Sheets {SPREADSHEET_ID}")
    else:
        print(f"[WARN] Ошибка загрузки листа '{version}' в Google Sheets")

    return out_path

# ---------- Планирование ----------
def next_run_datetime(now: datetime) -> datetime:
    for delta in range(0, 8):
        dt = now + timedelta(days=delta)
        for t in OFFICIAL_SLOTS:
            candidate = dt.replace(hour=t.hour, minute=t.minute, second=0, microsecond=0)
            if candidate > now:
                return candidate
    return now + timedelta(hours=1)

def main():
    print("=== Питон скрипт для Allure SWAT 733 ===")
    version = input("Введите номер релиза (например 7.3.7000): ").strip()
    if not version:
        print("[ERROR] Не указан номер релиза.")
        sys.exit(2)

    now = now_msk()
    nxt = next_run_datetime(now)
    wait_sec = max(1, int((nxt - now).total_seconds()))
    print(f"Следующий запуск (МСК) в {nxt.strftime('%Y-%m-%d %H:%M:%S')}, через {wait_sec} секунд.")
    print("Для ручного запуска введите цифру 1: ")

    manual_flag = {"v": False}
    import threading
    def _listen_manual():
        while True:
            try:
                s = input().strip()
            except EOFError:
                break
            if s == "1":
                manual_flag["v"] = True
    listener = threading.Thread(target=_listen_manual, daemon=True)
    listener.start()

    while True:
        now = now_msk()
        cur_label = time_to_label(dtime(now.hour, now.minute))
        if any(cur_label == time_to_label(t) for t in OFFICIAL_SLOTS):
            run_collect(version, manual_between=False)
            time.sleep(60)
            continue

        nxt = next_run_datetime(now)
        remaining = max(1, int((nxt - now).total_seconds()))
        sec_passed = 0
        while sec_passed < remaining:
            time.sleep(1)
            sec_passed += 1
            if manual_flag["v"]:
                run_collect(version, manual_between=True)
                manual_flag["v"] = False
                nxt = next_run_datetime(now_msk())
                remaining = max(1, int((nxt - now_msk()).total_seconds()))
                sec_passed = 0
                continue
            if sec_passed % 60 == 0 or sec_passed == remaining:
                left = remaining - sec_passed
                print(f"До следующего запуска {left} секунд.")

if __name__ == "__main__":
    main()
