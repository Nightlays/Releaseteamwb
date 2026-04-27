#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
import json
import re
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


try:
    from collect_uvu_testcases import ALLURE_API_TOKEN as DEFAULT_ALLURE_API_TOKEN
except Exception:
    DEFAULT_ALLURE_API_TOKEN = ""


BASE_URL = "https://allure-testops.wb.ru"
PROJECT_ID = 7
AUDIT_PAGE_SIZE = 1000
MAX_WORKERS = 12
REQUEST_TIMEOUT = 30
RETRIES = 3
RETRY_SLEEP_SEC = 1.0
UWU_FIELD_NAME = "UwU"

ALLOWED_LOGINS_RAW = """
kolosov.roman Роман Колосов - Старший смены
dolgov.viktor7 Виктор Долгов - Старший смены
egorov.valeriy15 Валерий Егоров - Профиль
kuklikov.andrey Андрей Кукликов - Карточка товара
pustovalova.g Галина Пустовалова -
musa.layla Лайла Муса - Core / Auth
geraskin.denis Денис Гераськин - Корзина (1ШК)
sergienko.e16 Екатерина Сергиенко - Корзина (2ШК)
moroz.tatyana8 Татьяна Мороз - Корзина (Курьер WB)
petrov.dmitriy177 Дмитрий Петров - Корзина (Заказы)
knyazev.pavel13 Павел Князев - Финтех (Частями)
deymos.krista Криста Деймос - Финтех (Недвижимость)
kalmykov.dmitriy11 Дмитрий Калмыков - Payments
konstantinova.e37 Екатерина Константинова - Travel.Avia
parshina.taisiya Таисия Паршина - Travel.Avia
zhvaev.nikita Никита Жваев - Карточка товара
dimitriev.david Давид Димитриев - Финтех (Спецпроекты)
borovkova.k7 Ксения Боровкова - Корзина (2ШК)
kapustin.dmitriy27 Дмитрий Капустин - Корзина (Заказы)
nazarkov.maksim Максим Назарков - Рекомендации
bahareva.a14 Анастасия Бахарева - Отзывы
deyneka.daniil Даниил Дейнека - Payments
zhashkov.egor Егор Жашков - Core
petrov.artemiy7 Артемий Петров - Корзина (Способы доставок)
dunaev.arseniy Арсений Дунаев - Корзина (2ШК)
lazarev.nikita52 Никита Лазарев - Профиль
tapehina.a Анастасия Тапехина - BX (Поиск и Каталог)
novikova.e187 Екатерина Новикова - Корзина (Преордер)
kazakov.egor21 Егор Казаков - Корзина (2ШК)
ostapenko.alena4 Алёна Остапенко - Корзина (Кроссбордер)
kocoeva.sabrina Сабрина Коцоева - BX (Поиск и Каталог)
mutalieva.f Фердовс Муталиева - Частями (Финтех)
mamaev.andrey17 Мамаев Андрей - Частями (Финтех)
rumyancev.artem19 Артём Румянцев - Спецпроекты (Финтех)
"""

_THREAD_LOCAL = threading.local()
_CFV_CACHE: Dict[int, "CFVInfo"] = {}
_CFV_CACHE_LOCK = threading.Lock()


@dataclass(frozen=True)
class SnapshotItem:
    test_case_id: int
    name: str
    uwu_raw: Optional[str]
    platform: Optional[str]


@dataclass(frozen=True)
class ChangedCase:
    test_case_id: int
    name: str
    old_uwu: Optional[str]
    new_uwu: Optional[str]
    platform: Optional[str]


@dataclass(frozen=True)
class CFVInfo:
    cfv_id: int
    field_name: str
    value_name: str


@dataclass(frozen=True)
class AuditEvent:
    timestamp_ms: int
    username: str
    old_values: Tuple[str, ...]
    new_values: Tuple[str, ...]
    audit_ids: Tuple[int, ...]


def parse_args() -> argparse.Namespace:
    script_dir = Path(__file__).resolve().parent
    default_old = script_dir / "uvu_testcases1.json"
    default_new = script_dir / "uvu_testcases.json"
    default_output = script_dir / f"uvu_changes_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    parser = argparse.ArgumentParser(
        description=(
            "Сравнивает два snapshot-файла UwU, находит изменившиеся кейсы, "
            "восстанавливает последнее изменение UwU через audit/cfv и сохраняет Excel."
        )
    )
    parser.add_argument("--old-file", default=str(default_old), help="Старый snapshot JSON")
    parser.add_argument("--new-file", default=str(default_new), help="Новый snapshot JSON")
    parser.add_argument("--output", default=str(default_output), help="Путь до выходного Excel-файла")
    parser.add_argument("--base-url", default=BASE_URL, help="Базовый URL Allure TestOps")
    parser.add_argument("--project-id", type=int, default=PROJECT_ID, help="ID проекта Allure")
    parser.add_argument("--workers", type=int, default=MAX_WORKERS, help="Количество потоков для audit")
    parser.add_argument("--page-size", type=int, default=AUDIT_PAGE_SIZE, help="Размер страницы audit API")
    parser.add_argument("--timeout", type=int, default=REQUEST_TIMEOUT, help="Таймаут HTTP-запросов")
    parser.add_argument(
        "--allure-api-token",
        default=DEFAULT_ALLURE_API_TOKEN,
        help="Api-Token для Allure.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Ограничить число изменившихся кейсов для отладки. 0 = без ограничения.",
    )
    return parser.parse_args()


def parse_allowed_logins(raw: str) -> Dict[str, str]:
    result: Dict[str, str] = {}
    for line in raw.splitlines():
        line = line.strip()
        if not line:
            continue
        parts = line.split(maxsplit=1)
        login = parts[0].strip()
        label = parts[1].strip() if len(parts) > 1 else login
        employee_name = re.split(r"\s+-\s*", label, maxsplit=1)[0].strip()
        result[login] = employee_name or label
    return result


ALLOWED_LOGINS = parse_allowed_logins(ALLOWED_LOGINS_RAW)


def normalize_uwu(value: Optional[Any]) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    candidate = text.replace(",", ".")
    try:
        number = Decimal(candidate)
    except InvalidOperation:
        return text
    normalized = format(number.normalize(), "f")
    normalized = normalized.rstrip("0").rstrip(".")
    return normalized or "0"


def unique_preserve(items: Iterable[str]) -> Tuple[str, ...]:
    seen = set()
    out: List[str] = []
    for item in items:
        item = str(item).strip()
        if not item or item in seen:
            continue
        seen.add(item)
        out.append(item)
    return tuple(out)


def load_snapshot(path: Path) -> Dict[int, SnapshotItem]:
    data = json.loads(path.read_text(encoding="utf-8"))
    items = data.get("items") if isinstance(data, dict) else None
    if not isinstance(items, list):
        raise ValueError(f"В файле {path} нет массива items")

    result: Dict[int, SnapshotItem] = {}
    for item in items:
        if not isinstance(item, dict):
            continue
        raw_id = item.get("testCaseId")
        if raw_id is None:
            continue
        try:
            test_case_id = int(raw_id)
        except Exception:
            continue
        result[test_case_id] = SnapshotItem(
            test_case_id=test_case_id,
            name=str(item.get("name") or "").strip(),
            uwu_raw=item.get("uwuRaw"),
            platform=item.get("platform"),
        )
    return result


def find_changed_cases(old_items: Dict[int, SnapshotItem], new_items: Dict[int, SnapshotItem]) -> List[ChangedCase]:
    changed: List[ChangedCase] = []
    common_ids = sorted(set(old_items) & set(new_items))
    for test_case_id in common_ids:
        old_item = old_items[test_case_id]
        new_item = new_items[test_case_id]
        if normalize_uwu(old_item.uwu_raw) == normalize_uwu(new_item.uwu_raw):
            continue
        changed.append(
            ChangedCase(
                test_case_id=test_case_id,
                name=new_item.name or old_item.name,
                old_uwu=old_item.uwu_raw,
                new_uwu=new_item.uwu_raw,
                platform=new_item.platform or old_item.platform,
            )
        )
    return changed


def build_headers(api_token: str) -> Dict[str, str]:
    headers = {
        "Accept": "application/json",
        "User-Agent": "uvu-changes-report/1.0",
    }
    token = api_token.strip()
    if not token:
        raise RuntimeError("Пустой Api-Token для Allure.")
    if token.lower().startswith("api-token ") or token.lower().startswith("bearer "):
        headers["Authorization"] = token
    else:
        headers["Authorization"] = f"Api-Token {token}"
    return headers


def get_session(headers: Dict[str, str]) -> requests.Session:
    session = getattr(_THREAD_LOCAL, "session", None)
    if session is None:
        session = requests.Session()
        session.headers.update(headers)
        _THREAD_LOCAL.session = session
    return session


def http_get_json(
    headers: Dict[str, str],
    base_url: str,
    path: str,
    *,
    params: Optional[Dict[str, Any]] = None,
    referer: Optional[str] = None,
    timeout: int,
) -> Any:
    session = get_session(headers)
    url = f"{base_url.rstrip('/')}{path}"
    req_headers: Dict[str, str] = {}
    if referer:
        req_headers["Referer"] = referer

    last_error: Optional[Exception] = None
    for attempt in range(1, RETRIES + 1):
        try:
            response = session.get(url, params=params, headers=req_headers or None, timeout=timeout)
            if response.status_code >= 400:
                snippet = response.text[:500].strip()
                raise requests.HTTPError(
                    f"{response.status_code} {response.reason} for {url}. Response: {snippet}",
                    response=response,
                )
            return response.json()
        except Exception as exc:
            last_error = exc
            if attempt < RETRIES:
                time.sleep(RETRY_SLEEP_SEC * attempt)
                continue
            raise RuntimeError(str(last_error)) from last_error
    raise RuntimeError("HTTP request failed without explicit exception")


def load_audit_entries(
    headers: Dict[str, str],
    base_url: str,
    project_id: int,
    test_case_id: int,
    page_size: int,
    timeout: int,
) -> List[Dict[str, Any]]:
    entries: List[Dict[str, Any]] = []
    page = 0
    referer = f"{base_url.rstrip('/')}/project/{project_id}/test-cases/{test_case_id}/changelog?treeId=0"

    while True:
        payload = http_get_json(
            headers,
            base_url,
            "/api/testcase/audit",
            params={"testCaseId": test_case_id, "page": page, "size": page_size},
            referer=referer,
            timeout=timeout,
        )
        content = payload.get("content") if isinstance(payload, dict) else None
        if not isinstance(content, list) or not content:
            break
        entries.extend(item for item in content if isinstance(item, dict))

        if payload.get("last") is True:
            break
        total_pages = payload.get("totalPages")
        if isinstance(total_pages, int) and page >= total_pages - 1:
            break
        page += 1

    return entries


def resolve_cfv(
    headers: Dict[str, str],
    base_url: str,
    project_id: int,
    test_case_id: int,
    cfv_id: Any,
    timeout: int,
) -> Optional[CFVInfo]:
    try:
        normalized_id = int(cfv_id)
    except Exception:
        return None

    with _CFV_CACHE_LOCK:
        cached = _CFV_CACHE.get(normalized_id)
    if cached is not None:
        return cached

    referer = f"{base_url.rstrip('/')}/project/{project_id}/test-cases/{test_case_id}/changelog?treeId=0"
    payload = http_get_json(
        headers,
        base_url,
        f"/api/cfv/{normalized_id}",
        referer=referer,
        timeout=timeout,
    )

    custom_field = payload.get("customField") if isinstance(payload, dict) else None
    field_name = ""
    if isinstance(custom_field, dict):
        field_name = str(custom_field.get("name") or "").strip()
    value_name = str(payload.get("name") or "").strip() if isinstance(payload, dict) else ""
    info = CFVInfo(cfv_id=normalized_id, field_name=field_name, value_name=value_name)

    with _CFV_CACHE_LOCK:
        _CFV_CACHE[normalized_id] = info
    return info


def extract_uwu_values_from_ids(
    headers: Dict[str, str],
    base_url: str,
    project_id: int,
    test_case_id: int,
    ids: Any,
    timeout: int,
) -> Tuple[str, ...]:
    if not isinstance(ids, list):
        return tuple()
    values: List[str] = []
    for cfv_id in ids:
        info = resolve_cfv(headers, base_url, project_id, test_case_id, cfv_id, timeout)
        if info is None or info.field_name != UWU_FIELD_NAME:
            continue
        if info.value_name:
            values.append(info.value_name)
    return unique_preserve(values)


def build_latest_uwu_event(
    headers: Dict[str, str],
    base_url: str,
    project_id: int,
    test_case_id: int,
    audit_entries: Sequence[Dict[str, Any]],
    timeout: int,
) -> Optional[AuditEvent]:
    grouped: Dict[Tuple[int, str], Dict[str, Any]] = {}

    for entry in audit_entries:
        timestamp_ms = entry.get("timestamp")
        username = str(entry.get("username") or "").strip()
        if not isinstance(timestamp_ms, int) or not username:
            continue

        data_items = entry.get("data")
        if not isinstance(data_items, list):
            continue

        old_values: List[str] = []
        new_values: List[str] = []

        for data_item in data_items:
            if not isinstance(data_item, dict):
                continue
            if str(data_item.get("type") or "").strip() != "test_case_custom_field":
                continue
            diff = data_item.get("diff")
            if not isinstance(diff, dict):
                continue
            ids_block = diff.get("ids")
            if not isinstance(ids_block, dict):
                continue
            old_values.extend(
                extract_uwu_values_from_ids(
                    headers,
                    base_url,
                    project_id,
                    test_case_id,
                    ids_block.get("oldValue"),
                    timeout,
                )
            )
            new_values.extend(
                extract_uwu_values_from_ids(
                    headers,
                    base_url,
                    project_id,
                    test_case_id,
                    ids_block.get("newValue"),
                    timeout,
                )
            )

        old_values_tuple = unique_preserve(old_values)
        new_values_tuple = unique_preserve(new_values)
        if not old_values_tuple and not new_values_tuple:
            continue

        key = (timestamp_ms, username)
        bucket = grouped.setdefault(
            key,
            {
                "timestamp_ms": timestamp_ms,
                "username": username,
                "old_values": [],
                "new_values": [],
                "audit_ids": [],
            },
        )
        bucket["old_values"].extend(old_values_tuple)
        bucket["new_values"].extend(new_values_tuple)
        raw_audit_id = entry.get("id")
        if isinstance(raw_audit_id, int):
            bucket["audit_ids"].append(raw_audit_id)

    events: List[AuditEvent] = []
    for bucket in grouped.values():
        events.append(
            AuditEvent(
                timestamp_ms=bucket["timestamp_ms"],
                username=bucket["username"],
                old_values=unique_preserve(bucket["old_values"]),
                new_values=unique_preserve(bucket["new_values"]),
                audit_ids=tuple(sorted(set(bucket["audit_ids"]), reverse=True)),
            )
        )

    if not events:
        return None

    events.sort(key=lambda event: (event.timestamp_ms, max(event.audit_ids or (0,))), reverse=True)
    return events[0]


def format_timestamp_msk(timestamp_ms: int) -> str:
    tz = timezone(timedelta(hours=3))
    dt = datetime.fromtimestamp(timestamp_ms / 1000, tz=tz)
    return dt.strftime("%Y-%m-%d %H:%M:%S")


def join_values(values: Sequence[str]) -> str:
    cleaned = [str(value).strip() for value in values if str(value).strip()]
    return " | ".join(cleaned)


def parse_numeric_value(raw: Optional[Any]) -> Optional[float]:
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None
    try:
        return float(text.replace(",", "."))
    except ValueError:
        return None


def first_non_empty(values: Sequence[str], fallback: Optional[Any]) -> str:
    for value in values:
        text = str(value).strip()
        if text:
            return text
    if fallback is None:
        return ""
    return str(fallback).strip()


def analyze_case(
    headers: Dict[str, str],
    base_url: str,
    project_id: int,
    changed_case: ChangedCase,
    page_size: int,
    timeout: int,
) -> Dict[str, Any]:
    entries = load_audit_entries(headers, base_url, project_id, changed_case.test_case_id, page_size, timeout)
    latest_event = build_latest_uwu_event(
        headers,
        base_url,
        project_id,
        changed_case.test_case_id,
        entries,
        timeout,
    )
    return {
        "changed_case": changed_case,
        "latest_event": latest_event,
        "audit_entries_count": len(entries),
    }


def autosize_worksheet(ws) -> None:
    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_len = 0
        for cell in column_cells:
            cell_value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(cell_value))
        ws.column_dimensions[column_letter].width = min(max(max_len + 2, 12), 60)


def write_excel(rows: Sequence[Dict[str, Any]], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "UwU changes"

    headers = [
        "Дата изменения (МСК)",
        "Тест кейс",
        "ФИО",
        "Было",
        "Стало",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="BDBDBD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=title)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row_index, row in enumerate(rows, start=2):
        values = [
            row["changed_at"],
            row["test_case_id"],
            row["employee_name"],
            row["old_value"],
            row["new_value"],
        ]
        ws.append(values)
        for column_index in range(1, len(headers) + 1):
            cell = ws.cell(row=row_index, column=column_index)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

        id_cell = ws.cell(row=row_index, column=2)
        id_cell.hyperlink = row["changelog_url"]
        id_cell.style = "Hyperlink"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize_worksheet(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def build_report_rows(
    results: Sequence[Dict[str, Any]],
    allowed_logins: Dict[str, str],
    base_url: str,
    project_id: int,
) -> Tuple[List[Dict[str, Any]], Dict[str, int]]:
    report_rows: List[Dict[str, Any]] = []
    counters = {
        "total_analyzed": len(results),
        "without_uwu_audit": 0,
        "not_in_whitelist": 0,
        "filtered_zero_growth": 0,
        "exported": 0,
    }

    for result in results:
        changed_case: ChangedCase = result["changed_case"]
        latest_event: Optional[AuditEvent] = result["latest_event"]
        if latest_event is None:
            counters["without_uwu_audit"] += 1
            continue

        label = allowed_logins.get(latest_event.username)
        if label is None:
            counters["not_in_whitelist"] += 1
            continue

        old_value = first_non_empty(latest_event.old_values, changed_case.old_uwu)
        new_value = first_non_empty(latest_event.new_values, changed_case.new_uwu)
        old_number = parse_numeric_value(old_value)
        new_number = parse_numeric_value(new_value)

        if old_number == 0 and new_number is not None and new_number <= 50:
            counters["filtered_zero_growth"] += 1
            continue

        diff_value = -1.0
        if old_number is not None and new_number is not None:
            diff_value = abs(new_number - old_number)

        changelog_url = (
            f"{base_url.rstrip('/')}/project/{project_id}/test-cases/"
            f"{changed_case.test_case_id}/changelog?treeId=0"
        )
        report_rows.append(
            {
                "changed_at": format_timestamp_msk(latest_event.timestamp_ms),
                "test_case_id": changed_case.test_case_id,
                "employee_name": label,
                "old_value": old_value,
                "new_value": new_value,
                "changelog_url": changelog_url,
                "sort_diff": diff_value,
                "sort_timestamp_ms": latest_event.timestamp_ms,
            }
        )

    report_rows.sort(
        key=lambda row: (
            row["sort_diff"],
            row["sort_timestamp_ms"],
            row["test_case_id"],
        ),
        reverse=True,
    )
    counters["exported"] = len(report_rows)
    return report_rows, counters


def main() -> None:
    args = parse_args()

    old_file = Path(args.old_file).resolve()
    new_file = Path(args.new_file).resolve()
    output_path = Path(args.output).resolve()

    if not old_file.exists():
        raise FileNotFoundError(f"Не найден old snapshot: {old_file}")
    if not new_file.exists():
        raise FileNotFoundError(f"Не найден new snapshot: {new_file}")

    headers = build_headers(args.allure_api_token)

    old_items = load_snapshot(old_file)
    new_items = load_snapshot(new_file)
    changed_cases = find_changed_cases(old_items, new_items)

    if args.limit > 0:
        changed_cases = changed_cases[: args.limit]

    print(f"Сравнение snapshot-ов: найдено {len(changed_cases)} кейсов с изменившимся UwU.")
    print(f"Белый список логинов: {len(ALLOWED_LOGINS)}.")

    results: List[Dict[str, Any]] = []
    errors: List[str] = []

    with ThreadPoolExecutor(max_workers=max(1, args.workers)) as pool:
        future_map = {
            pool.submit(
                analyze_case,
                headers,
                args.base_url,
                args.project_id,
                changed_case,
                args.page_size,
                args.timeout,
            ): changed_case
            for changed_case in changed_cases
        }

        processed = 0
        total = len(future_map)
        for future in as_completed(future_map):
            changed_case = future_map[future]
            try:
                results.append(future.result())
            except Exception as exc:
                errors.append(f"testCaseId={changed_case.test_case_id}: {exc}")
            processed += 1
            if total and (processed == total or processed % 25 == 0):
                print(f"Обработано audit: {processed}/{total}")

    if errors:
        print(f"Ошибок при загрузке audit/cfv: {len(errors)}")
        for error in errors[:20]:
            print(f"  - {error}")
        if len(errors) > 20:
            print(f"  ... и ещё {len(errors) - 20}")

    rows, counters = build_report_rows(results, ALLOWED_LOGINS, args.base_url, args.project_id)
    write_excel(rows, output_path)

    print(f"Экспортировано строк в Excel: {counters['exported']}")
    print(f"Без UwU в audit: {counters['without_uwu_audit']}")
    print(f"Отфильтровано не по whitelist: {counters['not_in_whitelist']}")
    print(f"Отфильтровано по правилу 0 -> <=50: {counters['filtered_zero_growth']}")
    print(f"Файл сохранён: {output_path}")


if __name__ == "__main__":
    main()
